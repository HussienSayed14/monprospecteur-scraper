"""
excel_uploader.py
─────────────────
Writes lead data to an Excel file matching the Google Sheet column structure.

Usage:
    from excel_uploader import write_leads_to_excel

    rows = [
        {
            "Scripte":          "...",
            "Contact Id":       "...",
            "First Name":       "...",
            # ... any columns you want to fill, leave the rest out
        },
        ...
    ]
    write_leads_to_excel(rows, output_path="output/leads.xlsx")
"""

from pathlib import Path
from datetime import datetime
import zoneinfo

TORONTO_TZ = zoneinfo.ZoneInfo("America/Toronto")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise ImportError("Run: pip install openpyxl")


# ── All column names exactly as in the Google Sheet ──────────────────────────
# Order matters — this is the column order in the Excel file.
# Pass only the keys you want to fill; the rest will be blank.

COLUMNS = [
    "Scripte",
    "Contact Id",
    "Scores",
    "Type",
    "Lead Source",
    "Reference Number",
    "Type Propriete",
    "Price",
    "Verification",
    "Numero lot",
    "Contact Owner.id",
    "Contact Owner",
    "First Name",
    "Last Name",
    "Email",
    "Date of Birth",
    "Company Name.id",
    "Company Name",
    "Mobile",
    "Phone",
    "Other Phone",
    "Home Phone",
    "Email Opt Out",
    "Tag",
    "Description",
    "Created By.id",
    "Created By",
    "Modified By.id",
    "Modified By",
    "Created Time",
    "Modified Time",
    "Contact Name",
    "Last Activity Time",
    "Unsubscribed Mode",
    "Unsubscribed Time",
    "Other Unit",
    "Other Street Number",
    "Other Street",
    "Other City",
    "Other Zip",
    "Other State",
    "Other Country",
    "Mailing Unit",
    "Mailing Street Number",
    "Mailing Street",
    "Mailing City",
    "Mailing Zip",
    "Mailing State",
    "Mailing Country",
    "Source motivation",
    "Equity",
    "Mortgage",
    "Analyse risque",
    "Picture 1",
    "Google Drive",
    "Direct Mail Date",
    "Direct Mail Number",
    "Visite",
    "Date Visite",
    "Call",
]

# ── Lead Source mapping ───────────────────────────────────────────────────────

LEAD_SOURCE_MAP = {
    "Succession":       "Succession",
    "Avis de 60 jours": "60Daysnotice",
    "Vente pour taxes": "Vpti",
}

# ── Reference Number counter ──────────────────────────────────────────────────
# Tracks how many leads have been scraped today within this run.
_ref_counter: dict = {}  # key: "YYYYMMDD" -> count

def _next_ref_number() -> str:
    """
    Generate: JLR + YYYYMMDD + zero-padded 2-digit sequence per day.
    Always uses TODAY's date in Toronto timezone — not the lead's published date.
    Example: JLR2026031401, JLR2026031402 ...
    Counter resets each Toronto calendar day.
    """
    d = datetime.now(TORONTO_TZ).strftime("%Y%m%d")
    _ref_counter[d] = _ref_counter.get(d, 0) + 1
    return f"JLR{d}{_ref_counter[d]:02d}"


# ── Address parsing helpers ───────────────────────────────────────────────────

import re as _re

def _format_postal_code(raw: str) -> str:
    """Uppercase with space after 3rd char: "h2b2j3" or "H2B2J3" -> "H2B 2J3" """
    if not raw:
        return ""
    clean = raw.upper().replace(" ", "")
    if len(clean) == 6:
        return f"{clean[:3]} {clean[3:]}"
    return clean

def _title_case_name(name: str) -> str:
    """
    Clean name capitalisation preserving accented characters (UTF-8).
    "BEAUPRÉ, CHANTAL" -> "Beaupré, Chantal"
    "MUNICIPALITE REGIONALE" -> "Municipalite Regionale"
    Uses str.title() which handles unicode/accents correctly in Python 3.
    """
    if not name:
        return ""
    # Ensure we are working with a proper unicode string
    if isinstance(name, bytes):
        name = name.decode("utf-8", errors="replace")
    name = name.strip()
    # Handle "LastName, FirstName" format
    if "," in name:
        parts = name.split(",", 1)
        return ", ".join(p.strip().title() for p in parts)
    return name.title()

def _parse_street_number(street: str) -> tuple[str, str]:
    """
    Split street string into (street_number, street_name).
    Handles:
      "10672 Avenue De Lorimier"   -> ("10672", "Avenue De Lorimier")
      "6630, rue Eugene-Achard"    -> ("6630", "rue Eugene-Achard")   # comma after number
      "rue Eugene-Achard"          -> ("", "rue Eugene-Achard")       # no number
    """
    s = street.strip()
    # Remove leading/trailing commas that sometimes appear: "6630, rue..."
    m = _re.match(r'^(\d+[\w-]*)[,\s]+(.*)', s)
    if m:
        return m.group(1).rstrip(",").strip(), m.group(2).strip()
    return "", s

def _parse_unit_from_street(street: str) -> tuple[str, str]:
    """
    Extract unit prefix if present, return (unit, remaining_street).
    Handles:
      "1110-1110a Chemin..."   -> unit="1110a",  street="1110 Chemin..."
      "101-456 Rue Test"       -> unit="101",    street="456 Rue Test"
      "456 Rue Test, Apt 3"   -> unit="3",       street="456 Rue Test"
      "6630, rue Test"         -> unit="",        street="6630, rue Test"
    """
    s = street.strip()

    # Pattern: "NUM-NUMa? Street" where second part starts with same digits + optional letter
    # e.g. "1110-1110a Chemin" → street_num=1110, unit=1110a
    m = _re.match(r'^(\d+)-(\d+[a-zA-Z]?)\s+(.*)', s)
    if m:
        first_num  = m.group(1)
        second_num = m.group(2)
        rest       = m.group(3).strip()
        # If second part is just digits (different number) it's unit-streetnum format
        if second_num.isdigit():
            return first_num, f"{second_num} {rest}".strip()
        else:
            # second part has a letter suffix — it's a unit variant like "1110a"
            return second_num, f"{first_num} {rest}".strip()

    # Pattern: suffix apt/suite/unit/# keyword
    m2 = _re.search(r'[\s,]+(apt|app|suite|unit|#)\s*(\w+)', s, _re.IGNORECASE)
    if m2:
        return m2.group(2), s[:m2.start()].strip()

    return "", s


def _get_district_value(district_info: list, label: str) -> str:
    """Get value from districtInfo array by label."""
    for item in (district_info or []):
        if item.get("label") == label:
            return item.get("value", "")
    return ""

def _get_district_number(district_info: list, label: str):
    """Get value_number from districtInfo array by label."""
    for item in (district_info or []):
        if item.get("label") == label:
            return item.get("value_number")
    return None

def _clean_price(raw_value: str) -> str:
    """
    "838 100,00" -> "838100"
    Remove spaces, remove decimal fraction (,xx at end).
    """
    import re
    v = raw_value.replace(" ", "").strip()
    # Remove fraction: ,00 or .00 at end
    v = re.sub(r'[,\.]\d+$', "", v)
    # Remove any remaining non-digit chars
    v = re.sub(r'[^0-9]', "", v)
    return v


def _vpti_reference_number(creancier_person: dict, published: str = "") -> str:
    """
    VPTI reference number format: VPTI {YEAR} {person_name of Creancier}
    Example: VPTI 2026 Municipalite Regionale De Comte De L'islet
    """
    try:
        year = datetime.fromisoformat(published.replace("Z", "+00:00")).astimezone(TORONTO_TZ).year
    except Exception:
        year = datetime.now(TORONTO_TZ).year

    person_name = (
        creancier_person.get("person_name")
        or creancier_person.get("lastName")
        or creancier_person.get("name")
        or ""
    ).strip()

    return f"VPTI {year} {person_name}" if person_name else f"VPTI {year}"


# ── Data cleaning / mapping ───────────────────────────────────────────────────

def clean_lead(list_doc: dict, detail_doc: dict = None) -> dict:
    """
    Map raw API data to the Excel column structure.

    Args:
        list_doc:   raw document from the list API
        detail_doc: raw document from GET /documents/{id}

    Returns:
        dict with keys matching COLUMNS — pass directly to write_leads_to_excel()
    """
    # Use detail_doc as primary source when available (richer data),
    # fall back to list_doc for any missing fields
    doc = detail_doc if detail_doc else list_doc

    doc_type  = list_doc.get("type", "")
    published = list_doc.get("publishedDate", "")
    try:
        pub_date = datetime.fromisoformat(published.replace("Z", "+00:00"))
    except Exception:
        pub_date = datetime.utcnow()

    # ── Address parsing ───────────────────────────────────────────────
    address_street  = doc.get("addressStreet") or ""
    address_city    = doc.get("addressCity") or ""
    address_zip_raw = doc.get("addressZipCode") or ""
    address_zip     = _format_postal_code(address_zip_raw)

    unit, street_without_unit = _parse_unit_from_street(address_street)
    street_num, street_name   = _parse_street_number(street_without_unit)

    # ── District info ─────────────────────────────────────────────────
    district_info = doc.get("districtInfo", [])
    price_raw     = _get_district_value(district_info, "Valeur de l'immeuble")
    price         = _clean_price(price_raw) if price_raw else ""

    # ── Lead type detection ──────────────────────────────────────────
    is_vpti = doc_type == "Vente pour taxes"

    # ── Parties array — check top level then inside acts ─────────────
    parties = doc.get("parties") or []
    if not parties:
        acts = doc.get("acts") or []
        if acts:
            parties = acts[0].get("parties") or []

    # ── Contact person logic ──────────────────────────────────────────
    # Succession/60Days: Legataire > Debiteur > second party > first party
    # VPTI:              Creancier (the municipality/creditor)
    contact_person = {}
    if is_vpti:
        # For VPTI take Creancier — the entity owed taxes
        for party in parties:
            if party.get("name", "").strip().lower() == "creancier":
                values = party.get("values") or []
                if values:
                    contact_person = values[0]
                    break
        # Fallback: first party
        if not contact_person and parties:
            values = parties[0].get("values") or []
            if values:
                contact_person = values[0]
    else:
        # Succession / 60Days: Legataire > Debiteur > second > first
        for target_name in ("Legataire", "Debiteur"):
            for party in parties:
                if party.get("name", "").strip().lower() == target_name.lower():
                    values = party.get("values") or []
                    if values:
                        contact_person = values[0]
                        break
            if contact_person:
                break
        if not contact_person and len(parties) >= 2:
            values = parties[1].get("values") or []
            if values:
                contact_person = values[0]
        if not contact_person and parties:
            values = parties[0].get("values") or []
            if values:
                contact_person = values[0]

    # ── First / Last Name ────────────────────────────────────────────
    # Succession/60Days: from contact_person (Legataire/Debiteur)
    # VPTI: from owners[0] — if no owner, "NF"
    owners = doc.get("owners") or []

    if is_vpti:
        if owners:
            first_owner = owners[0]
            raw_first   = (first_owner.get("firstName") or "").strip()
            raw_last    = (first_owner.get("lastName")  or first_owner.get("name") or "").strip()

            if raw_first:
                # Both first and last name present — use as is
                first_name = _title_case_name(raw_first)
                last_name  = _title_case_name(raw_last) if raw_last else "NF"
            elif raw_last:
                # Only a full name string (company or person) — split on first space
                # e.g. "10316555 Canada Inc." -> first="10316555", last="Canada Inc."
                # e.g. "Municipalite Regionale De Comte" -> first="Municipalite", last="Regionale De Comte"
                parts = raw_last.split(" ", 1)
                first_name = _title_case_name(parts[0])
                last_name  = _title_case_name(parts[1]) if len(parts) > 1 else "NF"
            else:
                first_name = "NF"
                last_name  = "NF"
        else:
            first_name = "NF"
            last_name  = "NF"
    else:
        first_name = _title_case_name(contact_person.get("firstName") or "")
        last_name  = _title_case_name(contact_person.get("lastName")  or "")

    # ── Mailing address ───────────────────────────────────────────────
    # VPTI:              from owners[0] — if no owner, all fields = "NF"
    # Succession/60Days: from contact_person (Legataire) — if no address, all = "NF"
    def _extract_mailing(person: dict) -> tuple:
        """Extract and parse mailing address from a person dict.
        Returns (unit, street_num, street_name, city, zip, has_data)."""
        street_raw = person.get("addressStreet") or ""
        city       = person.get("addressCity")   or ""
        zip_code   = _format_postal_code(person.get("addressZipCode") or "")
        unit, no_unit   = _parse_unit_from_street(street_raw)
        num, name       = _parse_street_number(no_unit)
        has_data = bool(street_raw or city or zip_code)
        return unit, num, name, city, zip_code, has_data

    def _nf_mailing():
        """Return address fields as NF — state and country are always Quebec/Canada."""
        return "NF", "NF", "NF", "NF", "NF", "Quebec", "Canada"

    if is_vpti:
        # VPTI: mailing comes from owners[0]
        # If no owners, or owner has no address data → address fields NF, state/country always set
        owner = owners[0] if owners else None
        if owner:
            m_unit, m_street_num, m_street_name, mailing_city, mailing_zip, has_mailing = _extract_mailing(owner)
            if has_mailing:
                mailing_state   = "Quebec"
                mailing_country = "Canada"
            else:
                m_unit, m_street_num, m_street_name, mailing_city, mailing_zip, mailing_state, mailing_country = _nf_mailing()
        else:
            m_unit, m_street_num, m_street_name, mailing_city, mailing_zip, mailing_state, mailing_country = _nf_mailing()
    else:
        # Succession/60Days: mailing comes from contact_person (Legataire)
        # If no address data → address fields NF, state/country always set
        m_unit, m_street_num, m_street_name, mailing_city, mailing_zip, has_mailing = _extract_mailing(contact_person)
        if has_mailing:
            mailing_state   = "Quebec"
            mailing_country = "Canada"
        else:
            m_unit, m_street_num, m_street_name, mailing_city, mailing_zip, mailing_state, mailing_country = _nf_mailing()

    # ── Cadastre (lot number) ─────────────────────────────────────────
    numero_lot = doc.get("cadastreNumber") or ""

    # ── Property type ─────────────────────────────────────────────────
    # VPTI: dynamic from propertyType field
    # Others: always "R-House"
    if is_vpti:
        type_propriete = doc.get("propertyType") or list_doc.get("propertyType") or "R-House"
    else:
        type_propriete = "R-House"

    row = {
        # Fixed values
        "Type":           "Prospect",
        "Type Propriete": type_propriete,

        # Mapped from doc type
        "Lead Source":    LEAD_SOURCE_MAP.get(doc_type, doc_type),

        # Reference Number:
        # Succession/60Days: JLR + YYYYMMDD + 2-digit sequence e.g. JLR2026031401
        # VPTI: VPTI {YEAR} {Creancier person_name}
        "Reference Number": _vpti_reference_number(contact_person, published) if is_vpti else _next_ref_number(),

        # Municipal assessment price (cleaned integer string)
        "Price":          price,

        # Cadastral number
        "Numero lot":     numero_lot,

        # Property address — "Other" fields
        "Other Unit":          unit,
        "Other Street Number": street_num,
        "Other Street":        street_name,
        "Other City":          address_city,
        "Other Zip":           address_zip,
        "Other State":         "Quebec",
        "Other Country":       "Canada",

        # Mailing address
        # Succession/60Days: from Legataire | VPTI: from owners[0] or NF
        "Mailing Unit":          m_unit,
        "Mailing Street Number": m_street_num,
        "Mailing Street":        m_street_name,
        "Mailing City":          mailing_city,
        "Mailing Zip":           mailing_zip,
        "Mailing State":         mailing_state,
        "Mailing Country":       mailing_country,

        # Contact name
        # Succession/60Days: Legataire | VPTI: Creancier
        "First Name":      first_name,
        "Last Name":       last_name,

        # Source motivation — filled from property_history API (passed in separately)
        "Source motivation": "",   # populated by caller after property_history API call

        # ── Still pending ──
        # "Mobile":       ...,
        # "Equity":       ...,
        # "Mortgage":     ...,
        # "Google Drive": ...,   <- filled by drive_uploader
    }

    return row


# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EVEN_FILL   = PatternFill("solid", fgColor="EEF2FF")
ODD_FILL    = PatternFill("solid", fgColor="FFFFFF")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="DDDDDD"),
)


# ── Main function ─────────────────────────────────────────────────────────────

def write_leads_to_excel(
    rows: list[dict],
    output_path: str = "output/leads.xlsx",
    sheet_name: str = "Leads",
) -> str:
    """
    Write a list of lead dicts to an Excel file.

    Each dict can contain any subset of the column names defined in COLUMNS.
    Keys not in COLUMNS are silently ignored.
    Columns not present in a row are left blank.

    Args:
        rows:        list of dicts, each dict is one lead row
        output_path: where to save the .xlsx file
        sheet_name:  worksheet name

    Returns:
        absolute path to the saved file
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── Header row ────────────────────────────────────────────────────
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=2):
        # Ensure proper unicode strings — preserves accented chars (é, à, ê, etc.)
        row_values = []
        for col in COLUMNS:
            val = row_data.get(col, "") or ""
            if isinstance(val, bytes):
                val = val.decode("utf-8", errors="replace")
            row_values.append(str(val) if val != "" else "")
        ws.append(row_values)

        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for cell in ws[row_idx]:
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = THIN_BORDER

    # ── Column widths ─────────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # Width = max of header length vs longest data value, capped at 50
        max_len = len(col_name)
        for row_data in rows:
            val = str(row_data.get(col_name, "") or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    wb.save(str(path))
    size_kb = path.stat().st_size / 1024
    print(f"✅ Excel saved → {path.resolve()}  ({len(rows)} rows, {size_kb:.1f} KB)")
    return str(path.resolve())


def append_leads_to_excel(
    rows: list[dict],
    output_path: str = "output/leads.xlsx",
    sheet_name: str = "Leads",
) -> str:
    """
    Append rows to an existing Excel file.
    If the file doesn't exist, creates it fresh.
    """
    path = Path(output_path)
    if not path.exists():
        return write_leads_to_excel(rows, output_path, sheet_name)

    wb = openpyxl.load_workbook(str(path))
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    next_row = ws.max_row + 1
    for row_idx, row_data in enumerate(rows, start=next_row):
        row_values = [row_data.get(col, "") for col in COLUMNS]
        ws.append(row_values)
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for cell in ws[row_idx]:
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")
            cell.border    = THIN_BORDER

    wb.save(str(path))
    print(f"✅ Appended {len(rows)} rows → {path.resolve()}")
    return str(path.resolve())


# ── Quick test ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    dummy_rows = [
        {
            "First Name":       "Denis",
            "Last Name":        "Lapalme",
            "Mailing Street":   "10672 Avenue De Lorimier",
            "Mailing City":     "Montreal",
            "Mailing Zip":      "H2B2J3",
            "Mailing State":    "QC",
            "Mailing Country":  "Canada",
            "Type Propriete":   "Unifamilial",
            "Type":             "Succession",
            "Lead Source":      "MonProspecteur",
            "Reference Number": "5eae5b61209a761738f4c776",
            "Google Drive":     "https://drive.google.com/...",
            "Created Time":     datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        },
        {
            "First Name":       "Marie",
            "Last Name":        "Tremblay",
            "Mailing Street":   "456 Rue Sainte-Catherine",
            "Mailing City":     "Montreal",
            "Mailing Zip":      "H3B1A7",
            "Type Propriete":   "Condo",
            "Type":             "Succession",
            "Lead Source":      "MonProspecteur",
            "Equity":           "120000",
            "Mortgage":         "80000",
        },
    ]

    write_leads_to_excel(dummy_rows, "output/test_leads.xlsx")